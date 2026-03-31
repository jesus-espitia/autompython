from flask import render_template, redirect, url_for, session, request, send_file
import pandas as pd
import os
import zipfile
from io import BytesIO

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
BASE_PUBLIC = os.path.join(PROJECT_ROOT, "archives", "public")



def buscar_header(ruta_excel, palabra_clave):
    df = pd.read_excel(ruta_excel, header=None)

    for i, row in df.iterrows():
        if palabra_clave in row.values:
            return i
    return None



def obtener_doliente(nombre):
    nombre = nombre.upper()

    if "AD-DC" in nombre:
        return "AD DS"
    elif "MIEMBROS" in nombre:
        return "INFRA"
    elif "APPWEB" in nombre:
        return "APPWEB"
    return "DESCONOCIDO"



def procesar_carpeta(ruta_carpeta, tipo):
    dfs = []

    for archivo in os.listdir(ruta_carpeta):
        if not archivo.endswith(".xlsx"):
            continue

        ruta = os.path.join(ruta_carpeta, archivo)

        palabra = "Node" if tipo == "cpu" else "Date"
        header_row = buscar_header(ruta, palabra)

        if header_row is None:
            continue

        df = pd.read_excel(ruta, header=header_row)


        df = df.loc[:, ~df.columns.isna()]


        df["Doliente"] = obtener_doliente(archivo)

        dfs.append(df)

    if dfs:
        return pd.concat(dfs, ignore_index=True)

    return pd.DataFrame()



def enriquecer_cpu_memoria(df_cpu):
    ruta_base = os.path.join(PROJECT_ROOT, "archives", "XLXS_IMPORT", "ALIMENTADORES")

    ruta_cpu = os.path.join(ruta_base, "RVTools_tabvCPU.xlsx")
    ruta_mem = os.path.join(ruta_base, "RVTools_tabvMemory.xlsx")
    ruta_fisicos = os.path.join(ruta_base, "Servidores_Fisicos.xlsx")

    # =========================
    # LEER ARCHIVOS
    # =========================
    df_cpu_tools = pd.read_excel(ruta_cpu)
    df_mem_tools = pd.read_excel(ruta_mem)
    df_fisicos = pd.read_excel(ruta_fisicos)

    # =========================
    # LIMPIAR COLUMNAS
    # =========================
    df_cpu_tools.columns = [col.strip() for col in df_cpu_tools.columns]
    df_mem_tools.columns = [col.strip() for col in df_mem_tools.columns]
    df_fisicos.columns = [col.strip() for col in df_fisicos.columns]

    # =========================
    # NORMALIZAR NOMBRES
    # =========================
    def limpiar_nombre(nombre):
        if pd.isna(nombre):
            return ""
        nombre = str(nombre).split("@")[0]
        return nombre.strip().lower()

    df_cpu["Node_clean"] = df_cpu["Node"].apply(limpiar_nombre)

    df_cpu_tools["VM_clean"] = df_cpu_tools["VM"].apply(limpiar_nombre)
    df_mem_tools["VM_clean"] = df_mem_tools["VM"].apply(limpiar_nombre)
    df_fisicos["Node_clean"] = df_fisicos["Node"].apply(limpiar_nombre)

    # =========================
    # MERGE CPU (RVTOOLS)
    # =========================
    df_cpu = df_cpu.merge(
        df_cpu_tools[["VM_clean", "CPUs", "Sockets", "Cores p/s"]],
        left_on="Node_clean",
        right_on="VM_clean",
        how="left"
    )

    # =========================
    # MERGE MEMORIA (RVTOOLS)
    # =========================
    if "SIZE GB" not in df_mem_tools.columns:
        raise Exception("No se encontró la columna 'SIZE GB' en RVTools_tabvMemory.xlsx")

    df_cpu = df_cpu.merge(
        df_mem_tools[["VM_clean", "SIZE GB"]],
        left_on="Node_clean",
        right_on="VM_clean",
        how="left"
    )

    # =========================
    # MERGE FISICOS
    # =========================
    df_cpu = df_cpu.merge(
        df_fisicos[["Node_clean", "CPUs", "SIZE GB", "Fisico"]],
        on="Node_clean",
        how="left",
        suffixes=("", "_fisico")
    )

    # =========================
    # COMPLETAR DATOS SI FALTAN
    # =========================
    df_cpu["CPUs"] = df_cpu["CPUs"].fillna(df_cpu["CPUs_fisico"])
    df_cpu["SIZE GB"] = df_cpu["SIZE GB"].fillna(df_cpu["SIZE GB_fisico"])

    # =========================
    # CREAR COLUMNA TIPO
    # =========================
    df_cpu["Tipo"] = df_cpu["Fisico"].apply(
        lambda x: "FISICO" if pd.notna(x) and str(x).strip() != "" else "VIRTUAL"
    )

    # =========================
    # LIMPIAR COLUMNAS AUX
    # =========================
    df_cpu.drop(columns=[
        "Node_clean",
        "VM_clean_x",
        "VM_clean_y",
        "CPUs_fisico",
        "SIZE GB_fisico",
        "Fisico"
    ], inplace=True, errors="ignore")

    # =========================
    # ORDENAR (Tipo al final)
    # =========================
    cols = [col for col in df_cpu.columns if col != "Tipo"] + ["Tipo"]
    df_cpu = df_cpu[cols]

    return df_cpu



def generar_zip(df_cpu, df_disco, mes):
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w") as z:

        # =========================
        # 📊 CPU CON FORMATO
        # =========================
        cpu_bytes = BytesIO()
        df_cpu.to_excel(cpu_bytes, index=False)

        # 🔥 Aplicar formato SIN alterar lógica existente
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        cpu_bytes.seek(0)
        wb = load_workbook(cpu_bytes)
        ws = wb.active

        # Buscar columna "Tipo"
        col_tipo = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Tipo":
                col_tipo = col
                break

        # Colores
        rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        # Aplicar colores
        if col_tipo:
            for row in range(2, ws.max_row + 1):
                valor = ws.cell(row=row, column=col_tipo).value

                if valor == "FISICO":
                    ws.cell(row=row, column=col_tipo).fill = rojo
                elif valor == "VIRTUAL":
                    ws.cell(row=row, column=col_tipo).fill = azul

        # Guardar Excel formateado
        cpu_final = BytesIO()
        wb.save(cpu_final)

        # 🔥 MISMO NOMBRE, MISMA LÓGICA
        z.writestr("CPU_Memoria_Unificado.xlsx", cpu_final.getvalue())

        # =========================
        # 📀 DISCO (NO SE TOCA)
        # =========================
        disco_bytes = BytesIO()
        df_disco.to_excel(disco_bytes, index=False)
        z.writestr("Disco_Unificado.xlsx", disco_bytes.getvalue())

    buffer.seek(0)
    return buffer



def F_discoMemoriaCpu():
    if "user" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        mes = request.form.get("mes")

        ruta_mes = os.path.join(BASE_PUBLIC, mes)
        ruta_cpu = os.path.join(ruta_mes, "CPU y Memoria")
        ruta_disco = os.path.join(ruta_mes, "Disco")

        df_cpu = procesar_carpeta(ruta_cpu, "cpu")

        df_cpu = enriquecer_cpu_memoria(df_cpu)

        df_disco = procesar_carpeta(ruta_disco, "disco")

        zip_buffer = generar_zip(df_cpu, df_disco, mes)

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"{mes}_unificado.zip",
            mimetype="application/zip"
        )

    meses = [
        carpeta for carpeta in os.listdir(BASE_PUBLIC)
        if os.path.isdir(os.path.join(BASE_PUBLIC, carpeta))
    ]



    return render_template("F_discoMemoriaCpu.html", meses=meses)