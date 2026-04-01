from routes import funciones, comparacion, archivos, discoMemoriaCpu,documentosPublicos, notificaciones
import glob
from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session
import secrets
import os
from datetime import datetime, timedelta
from config import Config
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
from werkzeug.utils import secure_filename

# =========================
# CONFIGURACIÓN INICIAL
# =========================

app = Flask(__name__, template_folder="Templates", static_folder="Static")
notificaciones.activar_notificaciones(app)
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, "archives")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
DB_FOLDER = os.path.join(UPLOAD_FOLDER, "XLXS_IMPORT")
DB_FILE = os.path.join(DB_FOLDER, "database.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DB_FOLDER, exist_ok=True)

# =========================
# INICIALIZAR EXCEL
# =========================

def init_database():
    if not os.path.exists(DB_FILE):
        wb = Workbook()

        ws_users = wb.active
        ws_users.title = "users"
        ws_users.append(["id", "fullname", "email", "password_hash", "verified"])

        ws_codes = wb.create_sheet("codes")
        ws_codes.append(["email", "code", "expiration"])

        wb.save(DB_FILE)
        wb.close()

init_database()

# =========================
# LIMPIEZA SILENCIOSA DE CACHE PYTHON
# =========================

def limpiar_cache_python():
    try:
        cache_folders = [
            os.path.join(PROJECT_ROOT, "__pycache__"),
            os.path.join(PROJECT_ROOT, "routes", "__pycache__")
        ]

        for folder in cache_folders:
            if os.path.exists(folder):
                for file in os.listdir(folder):
                    if file.endswith(".pyc"):
                        try:
                            os.remove(os.path.join(folder, file))
                        except:
                            pass
    except:
        pass


@app.before_request
def ejecutar_limpieza_cache():
    limpiar_cache_python()

#====================================================================================================================================
#====================================================================================================================================
# =========================
# FUNCIONES AUXILIARES
# =========================

def save_verification_code(email, code):
    wb = load_workbook(DB_FILE)
    ws = wb["codes"]

    expiration = datetime.now() + timedelta(minutes=10)
    ws.append([email, str(code), expiration.strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(DB_FILE)
    wb.close()

def send_verification_email(to_email, code, fullname):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "🔐 Verifica tu cuenta - PAMaap"
    msg["From"] = app.config["EMAIL_ADDRESS"]
    msg["To"] = to_email

    html_content = f"""
    <html>
    <body style="margin:0;padding:0;background:#050505;font-family:Segoe UI,Arial;">
        <table width="100%" cellpadding="0" cellspacing="0" style="padding:40px 0;">
            <tr>
                <td align="center">
                    <table width="500" cellpadding="0" cellspacing="0"
                        style="background:#0f0f0f;border-radius:12px;border:1px solid #252525;padding:40px;text-align:center;">
                        
                        <tr>
                            <td style="color:#ffffff;font-size:22px;font-weight:700;">
                                Bienvenido a PAMaap
                            </td>
                        </tr>

                        <tr>
                            <td style="color:#a0a0a0;font-size:14px;padding:20px 0;">
                                Hola <strong style="color:#ffffff;">{fullname}</strong>,<br>
                                Usa este código para verificar tu cuenta:
                            </td>
                        </tr>

                        <tr>
                            <td>
                                <div style="background:#141414;border:1px solid #252525;
                                    border-radius:10px;padding:20px;font-size:28px;
                                    letter-spacing:8px;font-weight:700;color:#ffd000;">
                                    {code}
                                </div>
                            </td>
                        </tr>

                        <tr>
                            <td style="color:#5a5a5a;font-size:12px;padding-top:20px;">
                                Este código expira en 10 minutos.
                            </td>
                        </tr>
                                <tr>
            <td align="center" style="padding-top:30px;">
                <img src="https://raw.githubusercontent.com/jesus-espitia/autompython/refs/heads/main/PAMaap/app/Static/img/firma/firma_PAMaap.png" alt="Firma PAMaap" style="width:590px;opacity:0.95;">
            </td>
        </tr>
                    </table>
                </td>
            </tr>
        </table>

    </body>
    </html>
    """

    msg.attach(MIMEText(html_content, "html"))

    try:
        server = smtplib.SMTP(app.config["SMTP_SERVER"], app.config["SMTP_PORT"])
        server.starttls()
        server.login(app.config["EMAIL_ADDRESS"], app.config["EMAIL_PASSWORD"])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print("Error enviando correo:", e)
        return False

# =========================
# LOGIN
# =========================

@app.route("/", methods=["GET", "POST"])
def login():

    if "user" in session:
        return redirect(url_for("home"))

    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        wb = load_workbook(DB_FILE)
        ws = wb["users"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, fullname, saved_email, password_hash, verified = row

            if saved_email == email:
                if not verified:
                    wb.close()
                    return render_template("index.html", message="Cuenta no verificada")

                if check_password_hash(password_hash, password):
                    session["user"] = {
                        "id": user_id,
                        "fullname": fullname,
                        "email": saved_email
                    }
                    wb.close()
                    return redirect(url_for("home"))

                wb.close()
                return render_template("index.html", message="Contraseña incorrecta")

        wb.close()
        return render_template("index.html", message="Usuario no encontrado")

    return render_template("index.html")

# =========================
# REGISTRO Y VERIFY 
# =========================

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":
        fullname = request.form["fullname"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:
            return render_template("register.html", message="Las contraseñas no coinciden")

        code = secrets.randbelow(900000) + 100000
        save_verification_code(email, code)

        if send_verification_email(email, code, fullname):
            session["temp_user"] = {
                "fullname": fullname,
                "email": email,
                "password": password
            }
            return redirect(url_for("verify"))

        return render_template("register.html", message="Error enviando correo")

    return render_template("register.html")

@app.route("/verify", methods=["GET", "POST"])
def verify():

    if "temp_user" not in session:
        return redirect(url_for("register"))

    temp_user = session["temp_user"]

    if request.method == "POST":
        user_code = request.form["code"]

        wb = load_workbook(DB_FILE)
        ws = wb["codes"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            saved_email, saved_code, expiration = row

            if saved_email == temp_user["email"] and str(saved_code) == str(user_code):

                if datetime.now() > datetime.strptime(expiration, "%Y-%m-%d %H:%M:%S"):
                    wb.close()
                    return render_template("verify.html", message="Código expirado")

                ws_users = wb["users"]
                user_id = ws_users.max_row
                hashed_password = generate_password_hash(temp_user["password"])

                ws_users.append([
                    user_id,
                    temp_user["fullname"],
                    temp_user["email"],
                    hashed_password,
                    True
                ])

                wb.save(DB_FILE)
                wb.close()

                session["user"] = {
                    "id": user_id,
                    "fullname": temp_user["fullname"],
                    "email": temp_user["email"]
                }

                session.pop("temp_user", None)
                return redirect(url_for("home"))

        wb.close()
        return render_template("verify.html", message="Código incorrecto")

    return render_template("verify.html")

#====================================================================================================================================
#====================================================================================================================================

# =========================
# DECORADOR DE SEGURIDAD
# =========================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# =========================
# CARPETA PRIVADA POR USUARIO
# =========================

def get_user_folder():
    user_id = session["user"]["id"]
    fullname = session["user"]["fullname"]

    safe_name = re.sub(r'[^a-zA-Z0-9]', '_', fullname).lower()
    folder_name = f"user_{user_id}_{safe_name}"

    user_folder = os.path.join(UPLOAD_FOLDER, folder_name)
    os.makedirs(user_folder, exist_ok=True)

    return user_folder

# =========================
# EXTENSIONES PERMITIDAS
# =========================

ALLOWED_EXTENSIONS = {
    # Documentos
    "pdf", "doc", "docx", "xls", "xlsx",
    "ppt", "pptx", "txt", "csv",

    # Imágenes
    "jpg", "jpeg", "png", "gif", "bmp", "webp"
}

def allowed_file(filename):
    if "." not in filename:
        return False
    extension = filename.rsplit(".", 1)[1].lower()
    return extension in ALLOWED_EXTENSIONS

# =========================
# DOCUMENTOS PRIVADOS
# =========================

@app.route("/documentos")
@login_required
def documentos():

    user_folder = get_user_folder()
    archivos = []

    for nombre in os.listdir(user_folder):
        ruta = os.path.join(user_folder, nombre)

        if os.path.isfile(ruta):
            archivos.append({
                "nombre": nombre,
                "tamano": round(os.path.getsize(ruta) / 1024, 2),
                "fecha": datetime.fromtimestamp(os.path.getctime(ruta)).strftime("%d %b %Y")
            })

    return render_template("documentos.html", archivos=archivos, user=session["user"])

@app.route("/ver/<nombre>")
@login_required
def ver_documento(nombre):
    return send_from_directory(get_user_folder(), nombre)

@app.route("/descargar/<nombre>")
@login_required
def descargar_documento(nombre):
    return send_from_directory(get_user_folder(), nombre, as_attachment=True)

@app.route("/eliminar/<nombre>", methods=["POST"])
@login_required
def eliminar_documento(nombre):
    ruta = os.path.join(get_user_folder(), nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
    return redirect(url_for("documentos"))

@app.route("/subir_documento", methods=["POST"])
@login_required
def subir_documento():

    if "archivo" not in request.files:
        return redirect(url_for("documentos"))

    archivo = request.files["archivo"]

    if archivo.filename == "":
        return redirect(url_for("documentos"))

    if not allowed_file(archivo.filename):
        return redirect(url_for("documentos"))

    user_folder = get_user_folder()


    filename = secure_filename(archivo.filename)

    ruta = os.path.join(user_folder, filename)

    archivo.save(ruta)

    return redirect(url_for("documentos"))
#====================================================================================================================================
#====================================================================================================================================
# =========================
#MI FICHA
# =========================

@app.route("/mi_ficha")
def mi_ficha():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template("mi_ficha.html")

# =========================
#configuracion
# =========================

@app.route("/configuracion")
def configuracion():
    if "user" not in session:
        return redirect(url_for("login"))
    return render_template("configuracion.html")

# =========================
#funciones
# =========================

@app.route("/funciones")
def funcionales():
    if "user" not in session:
        return redirect(url_for("login"))
    return funciones.funcion()
    

@app.route("/funcion/analisis")
def analisis():
    if "user" not in session:
        return redirect(url_for("login"))
    return funciones.F_analisis()

@app.route("/funcion/reportes", methods=["GET", "POST"])
def reportes():
    if "user" not in session:
        return redirect(url_for("login"))
    return funciones.F_reportes()

@app.route("/funcion/top5", methods=["GET", "POST"])
def top5():
    if "user" not in session:
        return redirect(url_for("login"))
    return funciones.F_top5()

@app.route("/funcion/comparacionArchivos", methods=["GET", "POST"])
def comparacionArchivo():
    if "user" not in session:
        return redirect(url_for("login"))
    return comparacion.F_comparacionArchivo()


@app.route("/funcion/archivoVisualizar", methods=["GET", "POST"])
def archivoVisualizar():
    if "user" not in session:
        return redirect(url_for("login"))
    return archivos.F_archivoVisualizar()

@app.route("/funcion/UnificarArchivos", methods=["GET", "POST"])
def discoMemoriaCpuVisualizar():
    if "user" not in session:
        return redirect(url_for("login"))
    return discoMemoriaCpu.F_discoMemoriaCpu()

@app.route("/documentosPublicos", methods=["GET", "POST"])
def documentosPublicosVisualizar():
    if "user" not in session:
        return redirect(url_for("login"))
    return documentosPublicos.documentosPublicos()

@app.route("/auditoria", methods=["GET", "POST"])
def prevencionNotificaciones():
    if "user" not in session:
        return redirect(url_for("login"))
    return notificaciones.generarNotificacionSMTP()
#====================================================================================================================================
#====================================================================================================================================

# =========================
# HOME Y LOGOUT
# =========================

@app.route("/home")
@login_required
def home():
    return render_template("home.html", user=session["user"])

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# =========================

if __name__ == "__main__":
    app.run(debug=True)