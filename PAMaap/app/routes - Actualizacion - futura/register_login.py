from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session
import secrets
import os
from datetime import datetime, timedelta
from config import Config
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder="Templates", static_folder="Static")
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, "archives")
DB_FOLDER = os.path.join(UPLOAD_FOLDER, "XLXS_IMPORT")
DB_FILE = os.path.join(DB_FOLDER, "database.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DB_FOLDER, exist_ok=True)

def init_database():
    if not os.path.exists(DB_FILE):
        wb = Workbook()

        ws_users = wb.active
        ws_users.title = "users"
        ws_users.append(["id", "fullname", "email", "password_hash", "verified"])

        ws_codes = wb.create_sheet("codes")
        ws_codes.append(["email", "code", "expiration"])

        wb.save(DB_FILE)
        wb.close()

init_database()
#====================================================================================================================================
#====================================================================================================================================

# =========================
# FUNCIONES AUXILIARES
# =========================

def save_verification_code(email, code):
    wb = load_workbook(DB_FILE)
    ws = wb["codes"]

    expiration = datetime.now() + timedelta(minutes=10)
    ws.append([email, str(code), expiration.strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(DB_FILE)
    wb.close()

def send_verification_email(to_email, code, fullname):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "🔐 Verifica tu cuenta - PAMaap"
    msg["From"] = app.config["EMAIL_ADDRESS"]
    msg["To"] = to_email

    html_content = f"""
    <html>
    <body style="margin:0;padding:0;background:#050505;font-family:Segoe UI,Arial;">
        <table width="100%" cellpadding="0" cellspacing="0" style="padding:40px 0;">
            <tr>
                <td align="center">
                    <table width="500" cellpadding="0" cellspacing="0"
                        style="background:#0f0f0f;border-radius:12px;border:1px solid #252525;padding:40px;text-align:center;">
                        
                        <tr>
                            <td style="color:#ffffff;font-size:22px;font-weight:700;">
                                Bienvenido a PAMaap
                            </td>
                        </tr>

                        <tr>
                            <td style="color:#a0a0a0;font-size:14px;padding:20px 0;">
                                Hola <strong style="color:#ffffff;">{fullname}</strong>,<br>
                                Usa este código para verificar tu cuenta:
                            </td>
                        </tr>

                        <tr>
                            <td>
                                <div style="background:#141414;border:1px solid #252525;
                                    border-radius:10px;padding:20px;font-size:28px;
                                    letter-spacing:8px;font-weight:700;color:#ffd000;">
                                    {code}
                                </div>
                            </td>
                        </tr>

                        <tr>
                            <td style="color:#5a5a5a;font-size:12px;padding-top:20px;">
                                Este código expira en 10 minutos.
                            </td>
                        </tr>

                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    msg.attach(MIMEText(html_content, "html"))

    try:
        server = smtplib.SMTP(app.config["SMTP_SERVER"], app.config["SMTP_PORT"])
        server.starttls()
        server.login(app.config["EMAIL_ADDRESS"], app.config["EMAIL_PASSWORD"])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print("Error enviando correo:", e)
        return False

# =========================
# LOGIN
# =========================

@app.route("/", methods=["GET", "POST"])
def login():

    if "user" in session:
        return redirect(url_for("home"))

    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        wb = load_workbook(DB_FILE)
        ws = wb["users"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, fullname, saved_email, password_hash, verified = row

            if saved_email == email:
                if not verified:
                    wb.close()
                    return render_template("index.html", message="Cuenta no verificada")

                if check_password_hash(password_hash, password):
                    session["user"] = {
                        "id": user_id,
                        "fullname": fullname,
                        "email": saved_email
                    }
                    wb.close()
                    return redirect(url_for("home"))

                wb.close()
                return render_template("index.html", message="Contraseña incorrecta")

        wb.close()
        return render_template("index.html", message="Usuario no encontrado")

    return render_template("index.html")

# =========================
# REGISTRO Y VERIFY 
# =========================

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":
        fullname = request.form["fullname"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:
            return render_template("register.html", message="Las contraseñas no coinciden")

        code = secrets.randbelow(900000) + 100000
        save_verification_code(email, code)

        if send_verification_email(email, code, fullname):
            session["temp_user"] = {
                "fullname": fullname,
                "email": email,
                "password": password
            }
            return redirect(url_for("verify"))

        return render_template("register.html", message="Error enviando correo")

    return render_template("register.html")

@app.route("/verify", methods=["GET", "POST"])
def verify():

    if "temp_user" not in session:
        return redirect(url_for("register"))

    temp_user = session["temp_user"]

    if request.method == "POST":
        user_code = request.form["code"]

        wb = load_workbook(DB_FILE)
        ws = wb["codes"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            saved_email, saved_code, expiration = row

            if saved_email == temp_user["email"] and str(saved_code) == str(user_code):

                if datetime.now() > datetime.strptime(expiration, "%Y-%m-%d %H:%M:%S"):
                    wb.close()
                    return render_template("verify.html", message="Código expirado")

                ws_users = wb["users"]
                user_id = ws_users.max_row
                hashed_password = generate_password_hash(temp_user["password"])

                ws_users.append([
                    user_id,
                    temp_user["fullname"],
                    temp_user["email"],
                    hashed_password,
                    True
                ])

                wb.save(DB_FILE)
                wb.close()

                session["user"] = {
                    "id": user_id,
                    "fullname": temp_user["fullname"],
                    "email": temp_user["email"]
                }

                session.pop("temp_user", None)
                return redirect(url_for("home"))

        wb.close()
        return render_template("verify.html", message="Código incorrecto")

    return render_template("verify.html")