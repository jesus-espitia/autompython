from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session
import secrets
import os
from datetime import datetime, timedelta
from config import Config
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder="Templates", static_folder="Static")
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, "archives")
DB_FOLDER = os.path.join(UPLOAD_FOLDER, "XLXS_IMPORT")
DB_FILE = os.path.join(DB_FOLDER, "database.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DB_FOLDER, exist_ok=True)

def init_database():
    if not os.path.exists(DB_FILE):
        wb = Workbook()

        ws_users = wb.active
        ws_users.title = "users"
        ws_users.append(["id", "fullname", "email", "password_hash", "verified"])

        ws_codes = wb.create_sheet("codes")
        ws_codes.append(["email", "code", "expiration"])

        wb.save(DB_FILE)
        wb.close()

init_database()
#====================================================================================================================================
#====================================================================================================================================

# =========================
# LOGIN
# =========================

@app.route("/", methods=["GET", "POST"])
def login():

    if "user" in session:
        return redirect(url_for("home"))

    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        wb = load_workbook(DB_FILE)
        ws = wb["users"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, fullname, saved_email, password_hash, verified = row

            if saved_email == email:
                if not verified:
                    wb.close()
                    return render_template("index.html", message="Cuenta no verificada")

                if check_password_hash(password_hash, password):
                    session["user"] = {
                        "id": user_id,
                        "fullname": fullname,
                        "email": saved_email
                    }
                    wb.close()
                    return redirect(url_for("home"))

                wb.close()
                return render_template("index.html", message="Contraseña incorrecta")

        wb.close()
        return render_template("index.html", message="Usuario no encontrado")

    return render_template("index.html")