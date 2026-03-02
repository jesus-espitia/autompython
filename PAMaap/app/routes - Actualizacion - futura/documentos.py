from flask import Flask, render_template, send_from_directory, request, redirect, url_for, session
import secrets
import os
from datetime import datetime, timedelta
from config import Config
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder="Templates", static_folder="Static")
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, "archives")
DB_FOLDER = os.path.join(UPLOAD_FOLDER, "XLXS_IMPORT")
DB_FILE = os.path.join(DB_FOLDER, "database.xlsx")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DB_FOLDER, exist_ok=True)

def init_database():
    if not os.path.exists(DB_FILE):
        wb = Workbook()

        ws_users = wb.active
        ws_users.title = "users"
        ws_users.append(["id", "fullname", "email", "password_hash", "verified"])

        ws_codes = wb.create_sheet("codes")
        ws_codes.append(["email", "code", "expiration"])

        wb.save(DB_FILE)
        wb.close()

init_database()
#====================================================================================================================================
#====================================================================================================================================


# =========================
# DECORADOR DE SEGURIDAD
# =========================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# =========================
# CARPETA PRIVADA POR USUARIO
# =========================

def get_user_folder():
    user_id = session["user"]["id"]
    fullname = session["user"]["fullname"]

    safe_name = re.sub(r'[^a-zA-Z0-9]', '_', fullname).lower()
    folder_name = f"user_{user_id}_{safe_name}"

    user_folder = os.path.join(UPLOAD_FOLDER, folder_name)
    os.makedirs(user_folder, exist_ok=True)

    return user_folder

# =========================
# EXTENSIONES PERMITIDAS
# =========================

ALLOWED_EXTENSIONS = {
    # Documentos
    "pdf", "doc", "docx", "xls", "xlsx",
    "ppt", "pptx", "txt",

    # Imágenes
    "jpg", "jpeg", "png", "gif", "bmp", "webp"
}

def allowed_file(filename):
    if "." not in filename:
        return False
    extension = filename.rsplit(".", 1)[1].lower()
    return extension in ALLOWED_EXTENSIONS

# =========================
# DOCUMENTOS PRIVADOS
# =========================

@app.route("/documentos")
@login_required
def documentos():

    user_folder = get_user_folder()
    archivos = []

    for nombre in os.listdir(user_folder):
        ruta = os.path.join(user_folder, nombre)

        if os.path.isfile(ruta):
            archivos.append({
                "nombre": nombre,
                "tamano": round(os.path.getsize(ruta) / 1024, 2),
                "fecha": datetime.fromtimestamp(os.path.getctime(ruta)).strftime("%d %b %Y")
            })

    return render_template("documentos.html", archivos=archivos, user=session["user"])

@app.route("/ver/<nombre>")
@login_required
def ver_documento(nombre):
    return send_from_directory(get_user_folder(), nombre)

@app.route("/descargar/<nombre>")
@login_required
def descargar_documento(nombre):
    return send_from_directory(get_user_folder(), nombre, as_attachment=True)

@app.route("/eliminar/<nombre>", methods=["POST"])
@login_required
def eliminar_documento(nombre):
    ruta = os.path.join(get_user_folder(), nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
    return redirect(url_for("documentos"))

@app.route("/subir_documento", methods=["POST"])
@login_required
def subir_documento():

    if "archivo" not in request.files:
        return redirect(url_for("documentos"))

    archivo = request.files["archivo"]

    if archivo.filename == "":
        return redirect(url_for("documentos"))

    if not allowed_file(archivo.filename):
        return redirect(url_for("documentos"))

    user_folder = get_user_folder()
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = secure_filename(archivo.filename)
    nombre_final = f"{fecha_actual}_{filename}"

    archivo.save(os.path.join(user_folder, nombre_final))

    return redirect(url_for("documentos"))