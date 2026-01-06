import pandas as pd
import os
import re
from dateutil import parser
from datetime import time
from openpyxl import load_workbook


def separar_fecha_hora_reporte_scaneo():
    try:
        print("\n🧩 Separación REAL de Fecha y Hora - Excel compatible")
        print("⚠️ Columna requerida: 'Reporte Scaneo'\n")

        # ==================================================
        # 📂 RUTAS RELATIVAS (SEGÚN TU PROYECTO)
        # ==================================================
        base_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(base_dir, "..", ".."))
        excel_dir = os.path.join(project_root, "Excel")

        if not os.path.exists(excel_dir):
            print("❌ No existe la carpeta Excel.")
            return

        # ==================================================
        # 📄 LISTAR ARCHIVOS EXCEL
        # ==================================================
        archivos_excel = [
            f for f in os.listdir(excel_dir)
            if f.lower().endswith((".xlsx", ".xls"))
        ]

        if not archivos_excel:
            print("❌ No hay archivos Excel.")
            return

        print("📑 Archivos disponibles:\n")
        for i, archivo in enumerate(archivos_excel, 1):
            print(f"{i}. {archivo}")

        try:
            opcion = int(input("\nSelecciona archivo a modificar: "))
            if opcion < 1 or opcion > len(archivos_excel):
                print("❌ Opción inválida.")
                return
        except ValueError:
            print("❌ Debes ingresar un número.")
            return

        ruta_archivo = os.path.join(excel_dir, archivos_excel[opcion - 1])
        print(f"\n✅ Archivo seleccionado: {os.path.basename(ruta_archivo)}")

        # ==================================================
        # 📊 CARGAR EXCEL
        # ==================================================
        df = pd.read_excel(ruta_archivo)
        columnas_originales = df.columns.tolist()

        df.columns = [
            re.sub(r"\s+", " ", col).strip().lower()
            for col in df.columns
        ]

        if "reporte scaneo" not in df.columns:
            print("❌ No se encontró la columna 'Reporte Scaneo'")
            print("📌 Columnas:", columnas_originales)
            return

        # ==================================================
        # ➕ CREAR COLUMNAS LIMPIAS
        # ==================================================
        df["fecha scaneo"] = pd.NaT
        df["hora scaneo"] = None

        # ==================================================
        # 🧠 EXTRACCIÓN CORRECTA
        # ==================================================
        def extraer(valor):
            if pd.isna(valor):
                return pd.NaT, None

            try:
                dt = parser.parse(str(valor), dayfirst=True, fuzzy=True)
                fecha = pd.Timestamp(dt.date())
                hora = time(dt.hour, dt.minute, dt.second)
                return fecha, hora
            except Exception:
                return pd.NaT, None

        df[["fecha scaneo", "hora scaneo"]] = df["reporte scaneo"].apply(
            lambda x: pd.Series(extraer(x))
        )

        # ==================================================
        # 💾 GUARDAR EXCEL
        # ==================================================
        df.to_excel(ruta_archivo, index=False)

        # ==================================================
        # 🎨 FORZAR FORMATO EXCEL (OPENPYXL)
        # ==================================================
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        col_fecha = df.columns.get_loc("fecha scaneo") + 1
        col_hora = df.columns.get_loc("hora scaneo") + 1

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_fecha).number_format = "DD/MM/YYYY"
            ws.cell(row=row, column=col_hora).number_format = "HH:MM"

        wb.save(ruta_archivo)

        print("\n✅ Proceso finalizado correctamente.")
        print("📌 Fecha y hora ahora son valores reales de Excel.")
        print("📌 Filtros, orden y formato funcionan sin doble clic.\n")

    except Exception as e:
        print(f"\n❌ Error inesperado: {e}\n")


if __name__ == "__main__":
    separar_fecha_hora_reporte_scaneo()
