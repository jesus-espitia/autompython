import pandas as pd
import os
import re
from datetime import datetime, time
from openpyxl import load_workbook
from collections import Counter


def separar_fecha_hora_columnas_dinamicas():
    try:
        print("\n🧩 Separación REAL de Fecha y Hora")
        print("⚠️ El formato se detecta por MAYORÍA y se normaliza a DD/MM/YYYY ⚠️\n")

        # ==================================================
        # 📂 RUTAS
        # ==================================================
        base_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(base_dir, "..", ".."))
        excel_dir = os.path.join(project_root, "Excel")

        # ==================================================
        # 📄 LISTAR ARCHIVOS
        # ==================================================
        archivos_excel = [
            f for f in os.listdir(excel_dir)
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
        ]

        print("📑 Archivos disponibles:\n")
        for i, f in enumerate(archivos_excel, 1):
            print(f"{i}. {f}")

        opcion = int(input("\nSelecciona archivo a modificar: "))
        ruta_archivo = os.path.join(excel_dir, archivos_excel[opcion - 1])

        # ==================================================
        # 📊 CARGAR ARCHIVO
        # ==================================================
        df = pd.read_excel(ruta_archivo)
        columnas_originales = df.columns.tolist()

        print("\n📌 Columnas disponibles:\n")
        for c in columnas_originales:
            print(f"- {c}")

        col_map = {
            re.sub(r"\s+", " ", c).strip().lower(): c
            for c in columnas_originales
        }

        # ==================================================
        # 🔁 SELECCIÓN DE COLUMNAS
        # ==================================================
        columnas = set()
        while True:
            col = input("\nCopia el nombre EXACTO de la columna o 0 para salir: \n").strip()
            if col == "0":
                break

            key = re.sub(r"\s+", " ", col).strip().lower()
            if key not in col_map:
                print("❌ La columna no existe.")
                continue

            columnas.add(key)
            print(f"✅ Columna agregada: {col_map[key]}")

        if not columnas:
            print("❌ No se seleccionaron columnas.")
            return

        # ==================================================
        # 🧠 DETECTAR FORMATO DOMINANTE
        # ==================================================
        def detectar_formato(valores):
            patrones = Counter()

            for v in valores.dropna():
                s = str(v).strip()

                if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", s):
                    patrones["YMD"] += 1
                elif re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}", s):
                    d, m, _ = re.split("[-/]", s[:10])
                    if int(d) > 12:
                        patrones["DMY"] += 1
                    elif int(m) > 12:
                        patrones["MDY"] += 1
                    else:
                        patrones["AMB"] += 1

            if not patrones:
                return None

            formato, cantidad = patrones.most_common(1)[0]
            return formato

        # ==================================================
        # 🧠 PARSEO SEGÚN FORMATO
        # ==================================================
        def parsear(valor, formato):
            if pd.isna(valor):
                return pd.NaT, None

            texto = str(valor).strip()
            fecha_txt = texto[:10]

            fmt_map = {
                "DMY": "%d/%m/%Y",
                "MDY": "%m/%d/%Y",
                "YMD": "%Y/%m/%d"
            }

            sep = "/" if "/" in fecha_txt else "-"
            fmt = fmt_map.get(formato)
            if not fmt:
                return pd.NaT, None

            fmt = fmt.replace("/", sep)

            try:
                dt = datetime.strptime(fecha_txt, fmt)
                fecha = pd.Timestamp(dt.date())

                hora = None
                if len(texto) > 10:
                    try:
                        h = datetime.strptime(texto[11:16], "%H:%M")
                        hora = time(h.hour, h.minute)
                    except Exception:
                        pass

                return fecha, hora
            except Exception:
                return pd.NaT, None

        # ==================================================
        # ⚙ PROCESAR COLUMNAS
        # ==================================================
        for key in columnas:
            col_real = col_map[key]

            formato_dominante = detectar_formato(df[col_real])
            if not formato_dominante:
                print(f"⚠️ No se pudo detectar formato para {col_real}. Se omite.")
                continue

            print(f"📌 Columna '{col_real}' → formato detectado: {formato_dominante}")

            df[f"{col_real}_fecha"] = pd.NaT
            df[f"{col_real}_hora"] = None

            df[[f"{col_real}_fecha", f"{col_real}_hora"]] = df[col_real].apply(
                lambda x: pd.Series(parsear(x, formato_dominante))
            )

        # ==================================================
        # 💾 GUARDAR
        # ==================================================
        df.to_excel(ruta_archivo, index=False)

        # ==================================================
        # 🎨 FORMATO EXCEL FINAL (DD/MM/YYYY)
        # ==================================================
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        for key in columnas:
            col_real = col_map[key]
            cf = df.columns.get_loc(f"{col_real}_fecha") + 1
            ch = df.columns.get_loc(f"{col_real}_hora") + 1

            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=cf).number_format = "DD/MM/YYYY"
                ws.cell(row=r, column=ch).number_format = "HH:MM"

        wb.save(ruta_archivo)

        print("\n✅ Proceso finalizado correctamente.")
        print("📌 Formato detectado por mayoría.")
        print("📌 Resultado final normalizado a DD/MM/YYYY.")

    except Exception as e:
        print(f"\n❌ Error: {e}")


if __name__ == "__main__":
    separar_fecha_hora_columnas_dinamicas()
